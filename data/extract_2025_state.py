#!/usr/bin/env python3
"""Extract 2025 state-level data from EIA sources and append to audit_all_data.csv.

Sources:
  - Capacity: EIA-860M Jan 2026 vintage, Operating Year == 2025
  - Retail prices: EIA Electric Power Monthly Table 5.06.B (Dec 2025 YTD = full-year 2025)
  - Wholesale/all-in/peak/queue: Inherited from parent ISO 2025 estimate rows in CSV

Usage: python3 data/extract_2025_state.py
"""

import csv
import os

import openpyxl

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(DATA_DIR, "audit_all_data.csv")
EIA860M_PATH = os.path.join(DATA_DIR, "eia860m", "january_generator2026.xlsx")
EPM_PATH = "/tmp/epm_table_b.xlsx"

# ---------------------------------------------------------------------------
# 32 states → parent ISO mapping
# ---------------------------------------------------------------------------

STATE_ISO = {
    "TX": "ERCOT",
    "OK": "SPP", "KS": "SPP", "NE": "SPP", "NM": "SPP",
    "IL": "MISO", "IN": "MISO", "MN": "MISO", "MI": "MISO",
    "IA": "MISO", "WI": "MISO", "LA": "MISO", "MS": "MISO",
    "MO": "MISO", "AR": "MISO", "KY": "MISO",
    "CA": "CAISO",
    "VA": "PJM", "PA": "PJM", "OH": "PJM", "NJ": "PJM",
    "MD": "PJM", "WV": "PJM", "NC": "PJM", "DE": "PJM",
    "NY": "NYISO",
    "MA": "ISO-NE", "CT": "ISO-NE", "ME": "ISO-NE",
    "NH": "ISO-NE", "VT": "ISO-NE", "RI": "ISO-NE",
}

# ---------------------------------------------------------------------------
# ELCC factors by technology category
# ---------------------------------------------------------------------------

# Generic ELCC factors
ELCC_GENERIC = {
    "gas": 0.95,
    "battery": 0.85,
    "solar": 0.325,   # midpoint of 30-35%
    "wind": 0.20,     # midpoint of 15-25%
    "other_firm": 0.80,
    "hydro": 0.50,
}

# ISO-specific overrides
ELCC_OVERRIDES = {
    "SPP":   {"battery": 0.90, "wind": 0.225},
    "CAISO": {"battery": 0.875},  # midpoint of 85-90%
    "MISO":  {"solar": 0.50},
}

# Map EIA-860M Technology strings to ELCC categories
TECH_CATEGORY = {
    "Solar Photovoltaic": "solar",
    "Batteries": "battery",
    "Onshore Wind Turbine": "wind",
    "Natural Gas Fired Combustion Turbine": "gas",
    "Natural Gas Fired Combined Cycle": "gas",
    "Natural Gas Internal Combustion Engine": "gas",
    "Petroleum Liquids": "gas",  # treat as firm dispatchable
    "Wood/Wood Waste Biomass": "other_firm",
    "Landfill Gas": "other_firm",
    "Conventional Hydroelectric": "hydro",
    "All Other": "other_firm",
}


def get_elcc_factor(tech_category, iso):
    """Get ELCC factor for a technology category within an ISO."""
    overrides = ELCC_OVERRIDES.get(iso, {})
    if tech_category in overrides:
        return overrides[tech_category]
    return ELCC_GENERIC.get(tech_category, 0.50)


# ---------------------------------------------------------------------------
# Parse EIA-860M for 2025 capacity by state
# ---------------------------------------------------------------------------

def parse_eia860m():
    """Parse EIA-860M Jan 2026 vintage for Operating Year == 2025.

    Returns dict: state_code -> {nameplate_mw, elcc_mw, project_count, tech_summary}
    """
    wb = openpyxl.load_workbook(EIA860M_PATH, read_only=True, data_only=True)
    ws = wb["Operating"]

    # Collect generators by state
    from collections import defaultdict
    state_generators = defaultdict(list)

    for row in ws.iter_rows(min_row=4, values_only=True):
        op_year = row[19]  # Operating Year
        if op_year != 2025:
            continue
        state = str(row[6]).strip() if row[6] else ""
        nameplate = row[12]  # Nameplate Capacity (MW)
        technology = str(row[15]).strip() if row[15] else "Unknown"

        if state not in STATE_ISO or not nameplate:
            continue
        try:
            mw = float(nameplate)
        except (ValueError, TypeError):
            continue

        state_generators[state].append({"mw": mw, "technology": technology})

    wb.close()

    # Compute nameplate, ELCC, project_count per state
    results = {}
    for state_code in STATE_ISO:
        gens = state_generators.get(state_code, [])
        nameplate_mw = sum(g["mw"] for g in gens)
        iso = STATE_ISO[state_code]

        elcc_mw = 0.0
        for g in gens:
            cat = TECH_CATEGORY.get(g["technology"], "other_firm")
            factor = get_elcc_factor(cat, iso)
            elcc_mw += g["mw"] * factor

        results[state_code] = {
            "nameplate_mw": round(nameplate_mw, 1),
            "elcc_mw": round(elcc_mw, 1),
            "project_count": len(gens),
        }

    return results


# ---------------------------------------------------------------------------
# Parse EPM Table 5.06.B for 2025 retail prices
# ---------------------------------------------------------------------------

# Map EIA state names to our 2-letter codes
EPM_STATE_NAMES = {
    "Texas": "TX", "Oklahoma": "OK", "Kansas": "KS", "Nebraska": "NE",
    "New Mexico": "NM", "Illinois": "IL", "Indiana": "IN",
    "Minnesota": "MN", "Michigan": "MI", "Iowa": "IA", "Wisconsin": "WI",
    "Louisiana": "LA", "Mississippi": "MS", "Missouri": "MO",
    "Arkansas": "AR", "Kentucky": "KY", "California": "CA",
    "Virginia": "VA", "Pennsylvania": "PA", "Ohio": "OH",
    "New Jersey": "NJ", "Maryland": "MD", "West Virginia": "WV",
    "North Carolina": "NC", "Delaware": "DE", "New York": "NY",
    "Massachusetts": "MA", "Connecticut": "CT", "Maine": "ME",
    "New Hampshire": "NH", "Vermont": "VT", "Rhode Island": "RI",
}


def parse_epm_retail_prices():
    """Parse EPM Table 5.06.B for 2025 All Sectors retail prices.

    Returns dict: state_code -> price_cents_kwh
    """
    wb = openpyxl.load_workbook(EPM_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Row 4 has headers: Col 0 = state, Col 9 = "All Sectors" Dec 2025 YTD
    prices = {}
    for row in ws.iter_rows(min_row=5, max_row=66, values_only=True):
        state_name = str(row[0]).strip() if row[0] else ""
        price_2025 = row[9]  # All Sectors, Dec 2025 YTD
        if state_name in EPM_STATE_NAMES and price_2025:
            try:
                prices[EPM_STATE_NAMES[state_name]] = round(float(price_2025), 2)
            except (ValueError, TypeError):
                pass

    wb.close()
    return prices


# ---------------------------------------------------------------------------
# Read existing CSV for ISO 2025 rows and 2024 state rows
# ---------------------------------------------------------------------------

def read_existing_csv():
    """Read audit_all_data.csv and return:
    - iso_2025: dict of ISO id -> row dict (for inherited fields)
    - state_2024: dict of state id -> row dict (for color_group, siting_regime, etc.)
    """
    with open(CSV_PATH, newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    iso_2025 = {}
    state_2024 = {}
    for row in rows:
        if row["view"] == "iso" and row["year"] == "2025":
            iso_2025[row["id"]] = row
        elif row["view"] == "state" and row["year"] == "2024":
            state_2024[row["id"]] = row

    return iso_2025, state_2024


# ---------------------------------------------------------------------------
# Build 32 new state rows
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    "view", "year", "id", "name", "region", "is_estimate", "color_group",
    "siting_regime", "confidence", "wholesale_price_mwh", "all_in_price_mwh",
    "retail_price_cents_kwh", "price_2023_mwh", "capacity_additions_mw",
    "capacity_additions_elcc_mw", "project_count", "peak_demand_gw",
    "queue_completion_pct", "queue_cohort", "avg_queue_duration_months",
    "qualitative_note", "source_price", "source_capacity", "source_peak",
    "source_queue",
]


def build_2025_state_rows(capacity, retail_prices, iso_2025, state_2024):
    """Build 32 CSV row dicts for (state, 2025)."""
    rows = []

    for state_code in sorted(STATE_ISO.keys()):
        iso_id = STATE_ISO[state_code]
        iso_row = iso_2025[iso_id]
        prev_row = state_2024[state_code]
        cap = capacity[state_code]
        retail = retail_prices.get(state_code, "")

        # Qualitative note: reuse 2024 note
        note = prev_row.get("qualitative_note", "")

        row = {
            "view": "state",
            "year": "2025",
            "id": state_code,
            "name": prev_row["name"],
            "region": prev_row["region"],
            "is_estimate": "True",
            "color_group": prev_row["color_group"],
            "siting_regime": prev_row.get("siting_regime", ""),
            "confidence": "estimated",
            "wholesale_price_mwh": iso_row["wholesale_price_mwh"],
            "all_in_price_mwh": iso_row["all_in_price_mwh"],
            "retail_price_cents_kwh": retail,
            "price_2023_mwh": "",
            "capacity_additions_mw": int(cap["nameplate_mw"]),
            "capacity_additions_elcc_mw": int(cap["elcc_mw"]),
            "project_count": cap["project_count"],
            "peak_demand_gw": prev_row["peak_demand_gw"],  # use 2024 as proxy
            "queue_completion_pct": prev_row["queue_completion_pct"],
            "queue_cohort": prev_row["queue_cohort"],
            "avg_queue_duration_months": prev_row["avg_queue_duration_months"],
            "qualitative_note": note,
            "source_price": f"Inherited from {iso_id} 2025 estimate",
            "source_capacity": "EIA-860M Jan 2026 vintage (operating year 2025)",
            "source_peak": "2024 proxy (EIA-861 state peak demand)",
            "source_queue": prev_row["source_queue"],
        }
        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Parsing EIA-860M Jan 2026 (operating year 2025)...")
    capacity = parse_eia860m()
    for sc in sorted(capacity.keys()):
        c = capacity[sc]
        print(f"  {sc}: {c['nameplate_mw']:.0f} MW nameplate, "
              f"{c['elcc_mw']:.0f} MW ELCC, {c['project_count']} generators")

    print("\nParsing EPM Table 5.06.B (2025 retail prices)...")
    retail_prices = parse_epm_retail_prices()
    for sc in sorted(retail_prices.keys()):
        print(f"  {sc}: {retail_prices[sc]} ¢/kWh")

    print("\nReading existing CSV...")
    iso_2025, state_2024 = read_existing_csv()
    print(f"  ISO 2025 rows: {len(iso_2025)}")
    print(f"  State 2024 rows: {len(state_2024)}")

    # Verify all 32 states have data
    missing_retail = [s for s in STATE_ISO if s not in retail_prices]
    if missing_retail:
        print(f"\n  WARNING: Missing retail prices for: {missing_retail}")

    print("\nBuilding 32 state rows...")
    new_rows = build_2025_state_rows(capacity, retail_prices, iso_2025, state_2024)

    # Append to CSV
    print(f"\nAppending {len(new_rows)} rows to {CSV_PATH}...")
    with open(CSV_PATH, "a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        for row in new_rows:
            writer.writerow(row)

    print("Done. Run 'npm run data:validate' to verify.")


if __name__ == "__main__":
    main()
